from __future__ import annotations

import hashlib
import io
import json
import re
import sqlite3
import unicodedata
import uuid
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from functools import lru_cache
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path
from typing import Any, Iterable

from langchain_core.documents import Document
from PIL import Image
from pypdf import PdfReader

from settings import (
    RAG_ALLOWED_EXTENSIONS,
    RAG_BATCH_MAX_BYTES,
    RAG_DOCUMENT_MAX_BYTES,
    RAG_IMAGE_MAX_BYTES,
    RAG_OCR_BACKEND,
    RAG_OCR_ENABLED,
    RAG_STATE_DB_PATH,
    RAG_UPLOAD_DIR,
)


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".xlsx",
    ".txt",
    ".md",
    *IMAGE_EXTENSIONS,
}
MIME_TYPES = {
    ".pdf": {"application/pdf"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/zip",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/zip",
    },
    ".txt": {"text/plain"},
    ".md": {"text/markdown", "text/x-markdown", "text/plain"},
    ".png": {"image/png"},
    ".jpg": {"image/jpeg", "image/pjpeg"},
    ".jpeg": {"image/jpeg", "image/pjpeg"},
    ".webp": {"image/webp"},
    ".bmp": {"image/bmp", "image/x-ms-bmp"},
    ".tif": {"image/tiff"},
    ".tiff": {"image/tiff"},
}
CANONICAL_MIME_TYPES = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".tif": "image/tiff",
    ".tiff": "image/tiff",
}
WINDOWS_RESERVED_NAMES = {
    "con",
    "prn",
    "aux",
    "nul",
    *(f"com{index}" for index in range(1, 10)),
    *(f"lpt{index}" for index in range(1, 10)),
}


class DocumentValidationError(ValueError):
    pass


class DocumentParseError(RuntimeError):
    pass


class _ClosingConnection(sqlite3.Connection):
    def __exit__(self, exc_type, exc_value, traceback):
        try:
            return super().__exit__(exc_type, exc_value, traceback)
        finally:
            self.close()


@dataclass(frozen=True)
class ParsedContent:
    normalized_text: str
    segments: list[dict[str, Any]]
    parser_name: str
    parser_version: str


@dataclass(frozen=True)
class UploadResult:
    status: str
    document_id: str
    version_id: str
    version_number: int
    original_name: str
    parse_status: str
    index_status: str
    message: str

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _package_version(package_name: str) -> str:
    try:
        return version(package_name)
    except PackageNotFoundError:
        return "unknown"


def normalize_text(text: str) -> str:
    if not isinstance(text, str):
        raise TypeError("Parsed content must be a Unicode string.")

    normalized = unicodedata.normalize("NFC", text)
    normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace("\u00a0", " ")
    normalized = "".join(
        character
        for character in normalized
        if character in {"\n", "\t"}
        or unicodedata.category(character) not in {"Cc", "Cf"}
    )

    lines: list[str] = []
    blank_line_seen = False
    for line in normalized.split("\n"):
        cleaned = line.rstrip()
        if not cleaned.strip():
            if lines and not blank_line_seen:
                lines.append("")
            blank_line_seen = True
            continue
        lines.append(cleaned)
        blank_line_seen = False
    return "\n".join(lines).strip()


def sanitize_filename(filename: str) -> str:
    normalized = unicodedata.normalize("NFC", str(filename or "")).strip()
    if (
        not normalized
        or "\x00" in normalized
        or "/" in normalized
        or "\\" in normalized
        or ":" in normalized
        or normalized in {".", ".."}
    ):
        raise DocumentValidationError("文件名无效或包含路径穿越字符。")

    cleaned = re.sub(r'[<>"/\\|?*\x00-\x1f]', "_", normalized)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    path = Path(cleaned)
    if not path.suffix or not path.stem:
        raise DocumentValidationError("文件名必须包含有效扩展名。")
    if path.stem.casefold() in WINDOWS_RESERVED_NAMES:
        raise DocumentValidationError("文件名使用了系统保留名称。")

    extension = path.suffix.lower()
    max_stem_length = max(1, 150 - len(extension))
    safe_name = f"{path.stem[:max_stem_length].rstrip(' .')}{extension}"
    if not safe_name or safe_name.startswith("."):
        raise DocumentValidationError("文件名清理后为空。")
    return safe_name


def decode_text_bytes(data: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "big5"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise DocumentParseError("文本编码无法识别，支持 UTF-8 和常见中文编码。")


def _safe_error_summary(error: Exception) -> str:
    message = f"{type(error).__name__}: {error}"
    message = re.sub(r"[\r\n\t]+", " ", message)
    message = re.sub(
        r"(?i)(api[_ -]?key|authorization|bearer|token)(\s*[:=]?\s*)\S+",
        r"\1\2[REDACTED]",
        message,
    )
    return message[:500]


def _verify_office_container(data: bytes, expected_prefix: str) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            names = archive.namelist()
    except (OSError, zipfile.BadZipFile) as exc:
        raise DocumentValidationError("Office 文件不是有效的 ZIP 容器。") from exc
    if not any(name.startswith(expected_prefix) for name in names):
        raise DocumentValidationError("文件内容与扩展名不匹配。")


def _verify_image(data: bytes, extension: str) -> None:
    expected_formats = {
        ".png": {"PNG"},
        ".jpg": {"JPEG"},
        ".jpeg": {"JPEG"},
        ".webp": {"WEBP"},
        ".bmp": {"BMP"},
        ".tif": {"TIFF"},
        ".tiff": {"TIFF"},
    }
    try:
        with Image.open(io.BytesIO(data)) as image:
            actual_format = str(image.format or "").upper()
            image.verify()
    except Exception as exc:
        raise DocumentValidationError("图片文件损坏或格式无法识别。") from exc
    if actual_format not in expected_formats[extension]:
        raise DocumentValidationError("图片内容与扩展名不匹配。")


def _validate_content_signature(data: bytes, extension: str) -> None:
    if extension == ".pdf":
        if not data.startswith(b"%PDF-"):
            raise DocumentValidationError("PDF 文件签名无效。")
    elif extension == ".docx":
        _verify_office_container(data, "word/")
    elif extension == ".xlsx":
        _verify_office_container(data, "xl/")
    elif extension in IMAGE_EXTENSIONS:
        _verify_image(data, extension)
    elif extension in {".txt", ".md"}:
        if b"\x00" in data:
            raise DocumentValidationError("文本文件包含二进制空字节。")
        decode_text_bytes(data)


def _segment(text: str, **metadata: Any) -> dict[str, Any] | None:
    normalized = normalize_text(text)
    if not normalized:
        return None
    return {"text": normalized, **metadata}


def _parse_pdf(data: bytes) -> ParsedContent:
    reader = PdfReader(io.BytesIO(data))
    segments = []
    for page_number, page in enumerate(reader.pages, start=1):
        item = _segment(page.extract_text() or "", page=page_number)
        if item:
            segments.append(item)
    return _parsed_result(segments, "pypdf", _package_version("pypdf"))


def _parse_docx(data: bytes) -> ParsedContent:
    from docx import Document as DocxDocument
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document = DocxDocument(io.BytesIO(data))
    segments: list[dict[str, Any]] = []
    paragraph_index = 0
    table_index = 0
    for child in document.element.body.iterchildren():
        if child.tag.endswith("}p"):
            paragraph_index += 1
            paragraph = Paragraph(child, document)
            text = paragraph.text.strip()
            style_name = str(getattr(paragraph.style, "name", "") or "")
            heading_match = re.search(r"(\d+)$", style_name)
            if text and style_name.lower().startswith("heading"):
                level = min(6, int(heading_match.group(1)) if heading_match else 1)
                text = f"{'#' * level} {text}"
            item = _segment(
                text,
                section_type="paragraph",
                paragraph=paragraph_index,
            )
        elif child.tag.endswith("}tbl"):
            table_index += 1
            table = Table(child, document)
            rows = [f"【表格 {table_index}】"]
            for row_number, row in enumerate(table.rows, start=1):
                cells = [
                    normalize_text(cell.text).replace("\n", " ")
                    for cell in row.cells
                ]
                rows.append(
                    f"Row {row_number}: "
                    + " | ".join(
                        f"C{cell_index}={value}"
                        for cell_index, value in enumerate(cells, start=1)
                    )
                )
            item = _segment(
                "\n".join(rows),
                section_type="table",
                table=table_index,
                row_start=1,
                row_end=len(table.rows),
            )
        else:
            item = None
        if item:
            segments.append(item)
    return _parsed_result(
        segments,
        "python-docx",
        _package_version("python-docx"),
    )


def _parse_xlsx(data: bytes) -> ParsedContent:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=False,
    )
    segments: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            lines = [f"## Sheet: {sheet.title}"]
            nonempty_rows = []
            for row_number, row in enumerate(sheet.iter_rows(), start=1):
                cells = []
                for column_number, cell in enumerate(row, start=1):
                    if cell.value is None:
                        continue
                    coordinate = f"{get_column_letter(column_number)}{row_number}"
                    value = normalize_text(str(cell.value)).replace("\n", " ")
                    cells.append(f"{coordinate}={value}")
                if cells:
                    nonempty_rows.append(row_number)
                    lines.append(f"Row {row_number}: " + " | ".join(cells))
            if nonempty_rows:
                item = _segment(
                    "\n".join(lines),
                    sheet=sheet.title,
                    row_start=min(nonempty_rows),
                    row_end=max(nonempty_rows),
                )
                if item:
                    segments.append(item)
    finally:
        workbook.close()
    return _parsed_result(
        segments,
        "openpyxl",
        _package_version("openpyxl"),
    )


def _parse_plain_text(data: bytes, *, markdown: bool) -> ParsedContent:
    text, encoding = decode_text_bytes(data)
    parser_name = "markdown_text" if markdown else "plain_text"
    item = _segment(text, encoding=encoding)
    return _parsed_result([item] if item else [], parser_name, "1")


@lru_cache(maxsize=1)
def _get_ocr_engine(backend: str):
    if backend != "rapidocr_onnxruntime":
        raise DocumentParseError(f"不支持的 OCR 后端：{backend}")
    try:
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as exc:
        raise DocumentParseError(
            "本地 OCR 未安装，请安装 rapidocr-onnxruntime。"
        ) from exc
    return RapidOCR()


def _parse_image(
    data: bytes,
    *,
    ocr_enabled: bool,
    ocr_backend: str,
) -> ParsedContent:
    if not ocr_enabled:
        raise DocumentParseError("图片 OCR 已关闭，图片不能进入知识库。")

    import numpy as np

    with Image.open(io.BytesIO(data)) as image:
        image_array = np.asarray(image.convert("RGB"))
    result, _ = _get_ocr_engine(ocr_backend)(image_array)
    if not result:
        raise DocumentParseError("OCR 未识别到可用文字。")

    lines = []
    scores = []
    for item in result:
        if len(item) < 3:
            continue
        text = normalize_text(str(item[1]))
        if text:
            lines.append(text)
            scores.append(float(item[2]))
    if not lines:
        raise DocumentParseError("OCR 返回结果中没有可用文字。")

    segment = _segment(
        "\n".join(lines),
        ocr_confidence=round(sum(scores) / len(scores), 4) if scores else None,
    )
    return _parsed_result(
        [segment] if segment else [],
        ocr_backend,
        _package_version("rapidocr-onnxruntime"),
    )


def _parsed_result(
    segments: Iterable[dict[str, Any]],
    parser_name: str,
    parser_version: str,
) -> ParsedContent:
    normalized_segments = [
        item
        for item in segments
        if item and normalize_text(str(item.get("text", "")))
    ]
    if not normalized_segments:
        raise DocumentParseError("解析完成但没有得到可索引文本。")
    normalized_text = normalize_text(
        "\n\n".join(str(item["text"]) for item in normalized_segments)
    )
    return ParsedContent(
        normalized_text=normalized_text,
        segments=normalized_segments,
        parser_name=parser_name,
        parser_version=parser_version,
    )


class DocumentService:
    def __init__(
        self,
        *,
        db_path: Path | str | None = None,
        upload_dir: Path | str | None = None,
        allowed_extensions: Iterable[str] | None = None,
        document_max_bytes: int = RAG_DOCUMENT_MAX_BYTES,
        image_max_bytes: int = RAG_IMAGE_MAX_BYTES,
        batch_max_bytes: int = RAG_BATCH_MAX_BYTES,
        ocr_enabled: bool = RAG_OCR_ENABLED,
        ocr_backend: str = RAG_OCR_BACKEND,
    ) -> None:
        self.db_path = Path(db_path or RAG_STATE_DB_PATH)
        self.upload_dir = Path(upload_dir or RAG_UPLOAD_DIR)
        configured = allowed_extensions or RAG_ALLOWED_EXTENSIONS
        self.allowed_extensions = tuple(
            dict.fromkeys(
                extension.lower()
                if str(extension).startswith(".")
                else f".{str(extension).lower()}"
                for extension in configured
            )
        )
        unsupported = set(self.allowed_extensions) - SUPPORTED_EXTENSIONS
        if unsupported:
            raise ValueError(f"不支持的上传扩展名配置：{sorted(unsupported)}")
        self.document_max_bytes = max(1, int(document_max_bytes))
        self.image_max_bytes = max(1, int(image_max_bytes))
        self.batch_max_bytes = max(1, int(batch_max_bytes))
        self.ocr_enabled = bool(ocr_enabled)
        self.ocr_backend = str(ocr_backend).strip().lower()
        self.initialize()

    def _connect(self) -> sqlite3.Connection:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(
            self.db_path,
            factory=_ClosingConnection,
        )
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        return connection

    def initialize(self) -> None:
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS rag_documents (
                    document_id TEXT PRIMARY KEY,
                    logical_name TEXT NOT NULL COLLATE NOCASE UNIQUE,
                    current_version INTEGER NOT NULL,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS rag_document_versions (
                    version_id TEXT PRIMARY KEY,
                    document_id TEXT NOT NULL,
                    version_number INTEGER NOT NULL,
                    original_name TEXT NOT NULL,
                    stored_path TEXT NOT NULL,
                    extension TEXT NOT NULL,
                    mime_type TEXT NOT NULL,
                    size_bytes INTEGER NOT NULL,
                    uploaded_by TEXT NOT NULL,
                    uploaded_at TEXT NOT NULL,
                    file_sha256 TEXT NOT NULL,
                    normalized_content_sha256 TEXT,
                    parser_name TEXT NOT NULL,
                    parser_version TEXT NOT NULL,
                    parse_status TEXT NOT NULL,
                    parse_error TEXT NOT NULL DEFAULT '',
                    index_status TEXT NOT NULL,
                    index_error TEXT NOT NULL DEFAULT '',
                    indexed_at TEXT,
                    chunk_count INTEGER NOT NULL DEFAULT 0,
                    normalized_text TEXT NOT NULL DEFAULT '',
                    segments_json TEXT NOT NULL DEFAULT '[]',
                    is_active INTEGER NOT NULL DEFAULT 0,
                    FOREIGN KEY(document_id) REFERENCES rag_documents(document_id),
                    UNIQUE(document_id, version_number)
                );

                CREATE UNIQUE INDEX IF NOT EXISTS uq_rag_version_file_sha256
                ON rag_document_versions(file_sha256);

                CREATE UNIQUE INDEX IF NOT EXISTS uq_rag_version_content_sha256
                ON rag_document_versions(normalized_content_sha256)
                WHERE normalized_content_sha256 IS NOT NULL
                  AND parse_status = 'parsed';

                CREATE INDEX IF NOT EXISTS ix_rag_versions_active
                ON rag_document_versions(parse_status, is_active, index_status);
                """
            )

    def validate_upload(
        self,
        filename: str,
        data: bytes,
        mime_type: str | None = None,
    ) -> tuple[str, str, str]:
        safe_name = sanitize_filename(filename)
        extension = Path(safe_name).suffix.lower()
        if extension not in self.allowed_extensions:
            raise DocumentValidationError(f"不允许的文件类型：{extension}")
        if not data:
            raise DocumentValidationError("上传文件为空。")

        size_limit = (
            self.image_max_bytes
            if extension in IMAGE_EXTENSIONS
            else self.document_max_bytes
        )
        if len(data) > size_limit:
            raise DocumentValidationError(
                f"文件超过大小限制：{len(data)} > {size_limit} bytes。"
            )

        declared_mime = str(mime_type or "").split(";", 1)[0].strip().lower()
        allowed_mimes = MIME_TYPES[extension]
        if (
            declared_mime
            and declared_mime != "application/octet-stream"
            and declared_mime not in allowed_mimes
        ):
            raise DocumentValidationError(
                f"MIME 与扩展名不匹配：{declared_mime} / {extension}"
            )
        _validate_content_signature(data, extension)
        return safe_name, extension, CANONICAL_MIME_TYPES[extension]

    def parse_content(self, extension: str, data: bytes) -> ParsedContent:
        if extension == ".pdf":
            return _parse_pdf(data)
        if extension == ".docx":
            return _parse_docx(data)
        if extension == ".xlsx":
            return _parse_xlsx(data)
        if extension == ".txt":
            return _parse_plain_text(data, markdown=False)
        if extension == ".md":
            return _parse_plain_text(data, markdown=True)
        if extension in IMAGE_EXTENSIONS:
            return _parse_image(
                data,
                ocr_enabled=self.ocr_enabled,
                ocr_backend=self.ocr_backend,
            )
        raise DocumentParseError(f"没有可用解析器：{extension}")

    def upload_many(
        self,
        files: Iterable[tuple[str, bytes, str | None]],
        *,
        uploaded_by: str,
    ) -> list[UploadResult]:
        pending = list(files)
        total_size = sum(len(data) for _, data, _ in pending)
        if total_size > self.batch_max_bytes:
            raise DocumentValidationError(
                f"批量上传总大小超过限制：{total_size} > "
                f"{self.batch_max_bytes} bytes。"
            )
        return [
            self.upload_file(
                filename,
                data,
                mime_type=mime_type,
                uploaded_by=uploaded_by,
            )
            for filename, data, mime_type in pending
        ]

    def upload_file(
        self,
        filename: str,
        data: bytes,
        *,
        mime_type: str | None = None,
        uploaded_by: str = "local_user",
    ) -> UploadResult:
        safe_name, extension, canonical_mime = self.validate_upload(
            filename,
            data,
            mime_type,
        )
        file_sha256 = hashlib.sha256(data).hexdigest()
        existing = self._find_version("file_sha256", file_sha256)
        if existing:
            return self._duplicate_result(existing, "duplicate_file")

        try:
            parsed = self.parse_content(extension, data)
        except Exception as exc:
            return self._record_parse_failure(
                safe_name=safe_name,
                extension=extension,
                mime_type=canonical_mime,
                data=data,
                file_sha256=file_sha256,
                uploaded_by=uploaded_by,
                error=exc,
            )

        content_sha256 = hashlib.sha256(
            parsed.normalized_text.encode("utf-8")
        ).hexdigest()
        existing = self._find_version(
            "normalized_content_sha256",
            content_sha256,
        )
        if existing:
            return self._duplicate_result(existing, "duplicate_content")

        return self._record_parsed_version(
            safe_name=safe_name,
            extension=extension,
            mime_type=canonical_mime,
            data=data,
            file_sha256=file_sha256,
            content_sha256=content_sha256,
            uploaded_by=uploaded_by,
            parsed=parsed,
        )

    def _find_version(
        self,
        field: str,
        value: str,
        *,
        connection: sqlite3.Connection | None = None,
    ) -> sqlite3.Row | None:
        if field not in {"file_sha256", "normalized_content_sha256"}:
            raise ValueError("Unsupported version lookup field.")
        owns_connection = connection is None
        active_connection = connection or self._connect()
        try:
            if field == "file_sha256":
                query = """
                SELECT v.*, d.logical_name
                FROM rag_document_versions AS v
                JOIN rag_documents AS d ON d.document_id = v.document_id
                WHERE v.file_sha256 = ?
                LIMIT 1
                """
            else:
                query = """
                SELECT v.*, d.logical_name
                FROM rag_document_versions AS v
                JOIN rag_documents AS d ON d.document_id = v.document_id
                WHERE v.normalized_content_sha256 = ?
                LIMIT 1
                """
            return active_connection.execute(query, (value,)).fetchone()
        finally:
            if owns_connection:
                active_connection.close()

    @staticmethod
    def _duplicate_result(row: sqlite3.Row, status: str) -> UploadResult:
        label = "原始文件重复" if status == "duplicate_file" else "规范化内容重复"
        return UploadResult(
            status=status,
            document_id=row["document_id"],
            version_id=row["version_id"],
            version_number=int(row["version_number"]),
            original_name=row["original_name"],
            parse_status=row["parse_status"],
            index_status=row["index_status"],
            message=f"{label}，复用现有文档版本，不重复保存或索引。",
        )

    def _next_version(
        self,
        connection: sqlite3.Connection,
        logical_name: str,
        *,
        active: bool,
    ) -> tuple[str, int, bool]:
        row = connection.execute(
            """
            SELECT document_id
            FROM rag_documents
            WHERE logical_name = ? COLLATE NOCASE
            """,
            (logical_name,),
        ).fetchone()
        if row:
            document_id = row["document_id"]
            version_number = int(
                connection.execute(
                    """
                    SELECT COALESCE(MAX(version_number), 0) + 1
                    FROM rag_document_versions
                    WHERE document_id = ?
                    """,
                    (document_id,),
                ).fetchone()[0]
            )
            return document_id, version_number, True

        document_id = uuid.uuid4().hex
        now = _utc_now()
        connection.execute(
            """
            INSERT INTO rag_documents (
                document_id, logical_name, current_version,
                is_active, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                document_id,
                logical_name,
                1,
                1 if active else 0,
                now,
                now,
            ),
        )
        return document_id, 1, False

    def _safe_storage_path(
        self,
        document_id: str,
        version_id: str,
        version_number: int,
        safe_name: str,
    ) -> Path:
        base = self.upload_dir.resolve()
        destination = (
            base
            / document_id
            / f"v{version_number:04d}"
            / f"{version_id}_{safe_name}"
        ).resolve()
        if base != destination and base not in destination.parents:
            raise DocumentValidationError("上传文件保存路径越界。")
        return destination

    def _write_original(
        self,
        document_id: str,
        version_id: str,
        version_number: int,
        safe_name: str,
        data: bytes,
    ) -> Path:
        destination = self._safe_storage_path(
            document_id,
            version_id,
            version_number,
            safe_name,
        )
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_bytes(data)
        return destination

    def _record_parsed_version(
        self,
        *,
        safe_name: str,
        extension: str,
        mime_type: str,
        data: bytes,
        file_sha256: str,
        content_sha256: str,
        uploaded_by: str,
        parsed: ParsedContent,
    ) -> UploadResult:
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            existing = self._find_version(
                "file_sha256",
                file_sha256,
                connection=connection,
            )
            if existing:
                return self._duplicate_result(existing, "duplicate_file")
            existing = self._find_version(
                "normalized_content_sha256",
                content_sha256,
                connection=connection,
            )
            if existing:
                return self._duplicate_result(existing, "duplicate_content")

            document_id, version_number, existed = self._next_version(
                connection,
                safe_name,
                active=True,
            )
            version_id = uuid.uuid4().hex
            stored_path = self._write_original(
                document_id,
                version_id,
                version_number,
                safe_name,
                data,
            )
            now = _utc_now()
            connection.execute(
                """
                UPDATE rag_document_versions
                SET is_active = 0,
                    index_status = CASE
                        WHEN index_status = 'indexed' OR chunk_count > 0
                        THEN 'pending_delete'
                        ELSE 'inactive'
                    END
                WHERE document_id = ? AND is_active = 1
                """,
                (document_id,),
            )
            connection.execute(
                """
                INSERT INTO rag_document_versions (
                    version_id, document_id, version_number, original_name,
                    stored_path, extension, mime_type, size_bytes,
                    uploaded_by, uploaded_at, file_sha256,
                    normalized_content_sha256, parser_name, parser_version,
                    parse_status, parse_error, index_status, index_error,
                    indexed_at, chunk_count, normalized_text,
                    segments_json, is_active
                ) VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'parsed', '',
                    'pending', '', NULL, 0, ?, ?, 1
                )
                """,
                (
                    version_id,
                    document_id,
                    version_number,
                    safe_name,
                    str(stored_path),
                    extension,
                    mime_type,
                    len(data),
                    str(uploaded_by or "local_user")[:100],
                    now,
                    file_sha256,
                    content_sha256,
                    parsed.parser_name,
                    parsed.parser_version,
                    parsed.normalized_text,
                    json.dumps(parsed.segments, ensure_ascii=False),
                ),
            )
            connection.execute(
                """
                UPDATE rag_documents
                SET current_version = ?, is_active = 1, updated_at = ?
                WHERE document_id = ?
                """,
                (version_number, now, document_id),
            )
        status = "new_version" if existed else "uploaded"
        return UploadResult(
            status=status,
            document_id=document_id,
            version_id=version_id,
            version_number=version_number,
            original_name=safe_name,
            parse_status="parsed",
            index_status="pending",
            message=(
                "已创建新版本，等待增量建库。"
                if existed
                else "文档已解析，等待增量建库。"
            ),
        )

    def _record_parse_failure(
        self,
        *,
        safe_name: str,
        extension: str,
        mime_type: str,
        data: bytes,
        file_sha256: str,
        uploaded_by: str,
        error: Exception,
    ) -> UploadResult:
        parser_names = {
            ".pdf": ("pypdf", _package_version("pypdf")),
            ".docx": ("python-docx", _package_version("python-docx")),
            ".xlsx": ("openpyxl", _package_version("openpyxl")),
            ".txt": ("plain_text", "1"),
            ".md": ("markdown_text", "1"),
        }
        parser_name, parser_version = parser_names.get(
            extension,
            (
                self.ocr_backend,
                _package_version("rapidocr-onnxruntime"),
            ),
        )
        parse_error = _safe_error_summary(error)
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            existing = self._find_version(
                "file_sha256",
                file_sha256,
                connection=connection,
            )
            if existing:
                return self._duplicate_result(existing, "duplicate_file")
            document_id, version_number, _ = self._next_version(
                connection,
                safe_name,
                active=False,
            )
            version_id = uuid.uuid4().hex
            stored_path = self._write_original(
                document_id,
                version_id,
                version_number,
                safe_name,
                data,
            )
            now = _utc_now()
            connection.execute(
                """
                INSERT INTO rag_document_versions (
                    version_id, document_id, version_number, original_name,
                    stored_path, extension, mime_type, size_bytes,
                    uploaded_by, uploaded_at, file_sha256,
                    normalized_content_sha256, parser_name, parser_version,
                    parse_status, parse_error, index_status, index_error,
                    indexed_at, chunk_count, normalized_text,
                    segments_json, is_active
                ) VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?,
                    'parse_failed', ?, 'not_indexed', '', NULL, 0, '', '[]', 0
                )
                """,
                (
                    version_id,
                    document_id,
                    version_number,
                    safe_name,
                    str(stored_path),
                    extension,
                    mime_type,
                    len(data),
                    str(uploaded_by or "local_user")[:100],
                    now,
                    file_sha256,
                    parser_name,
                    parser_version,
                    parse_error,
                ),
            )
        return UploadResult(
            status="parse_failed",
            document_id=document_id,
            version_id=version_id,
            version_number=version_number,
            original_name=safe_name,
            parse_status="parse_failed",
            index_status="not_indexed",
            message=parse_error,
        )

    def list_documents(self, *, include_inactive: bool = True) -> list[dict[str, Any]]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT d.*, v.version_id, v.original_name, v.size_bytes,
                       v.uploaded_by, v.uploaded_at, v.parse_status,
                       v.parse_error, v.index_status, v.index_error,
                       v.indexed_at, v.chunk_count
                FROM rag_documents AS d
                LEFT JOIN rag_document_versions AS v
                  ON v.document_id = d.document_id
                 AND v.version_number = d.current_version
                WHERE ? = 1 OR d.is_active = 1
                ORDER BY d.updated_at DESC, d.logical_name
                """,
                (1 if include_inactive else 0,),
            ).fetchall()
        return [dict(row) for row in rows]

    def list_versions(self, document_id: str | None = None) -> list[dict[str, Any]]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT v.*, d.logical_name, d.is_active AS document_active
                FROM rag_document_versions AS v
                JOIN rag_documents AS d ON d.document_id = v.document_id
                WHERE ? IS NULL OR v.document_id = ?
                ORDER BY v.uploaded_at DESC, v.version_number DESC
                """,
                (document_id, document_id),
            ).fetchall()
        return [dict(row) for row in rows]

    def deactivate_document(self, document_id: str) -> bool:
        now = _utc_now()
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            exists = connection.execute(
                "SELECT 1 FROM rag_documents WHERE document_id = ?",
                (document_id,),
            ).fetchone()
            if not exists:
                return False
            connection.execute(
                """
                UPDATE rag_documents
                SET is_active = 0, updated_at = ?
                WHERE document_id = ?
                """,
                (now, document_id),
            )
            connection.execute(
                """
                UPDATE rag_document_versions
                SET is_active = 0,
                    index_status = CASE
                        WHEN index_status = 'indexed' OR chunk_count > 0
                        THEN 'pending_delete'
                        ELSE 'inactive'
                    END
                WHERE document_id = ? AND is_active = 1
                """,
                (document_id,),
            )
        return True

    def load_active_documents(self) -> list[Document]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT v.*
                FROM rag_document_versions AS v
                JOIN rag_documents AS d ON d.document_id = v.document_id
                WHERE d.is_active = 1
                  AND v.is_active = 1
                  AND v.parse_status = 'parsed'
                ORDER BY v.uploaded_at, v.version_number
                """
            ).fetchall()

        documents: list[Document] = []
        for row in rows:
            try:
                segments = json.loads(row["segments_json"] or "[]")
            except json.JSONDecodeError:
                segments = []
            if not segments and row["normalized_text"]:
                segments = [{"text": row["normalized_text"]}]
            for segment in segments:
                text = normalize_text(str(segment.get("text", "")))
                if not text:
                    continue
                metadata = {
                    key: value
                    for key, value in segment.items()
                    if key != "text" and value is not None
                }
                metadata.update(
                    {
                        "document_id": row["document_id"],
                        "version_id": row["version_id"],
                        "version_number": int(row["version_number"]),
                        "source": (
                            f"upload/{row['document_id']}/"
                            f"{row['version_id']}/{row['original_name']}"
                        ),
                        "original_name": row["original_name"],
                        "parser_name": row["parser_name"],
                    }
                )
                documents.append(Document(page_content=text, metadata=metadata))
        return documents

    def active_version_ids(self) -> list[str]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT v.version_id
                FROM rag_document_versions AS v
                JOIN rag_documents AS d ON d.document_id = v.document_id
                WHERE d.is_active = 1
                  AND v.is_active = 1
                  AND v.parse_status = 'parsed'
                """
            ).fetchall()
        return [str(row["version_id"]) for row in rows]

    def mark_indexed(self, chunk_counts: dict[str, int]) -> None:
        now = _utc_now()
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            for version_id, chunk_count in chunk_counts.items():
                connection.execute(
                    """
                    UPDATE rag_document_versions
                    SET index_status = 'indexed', index_error = '',
                        indexed_at = ?, chunk_count = ?
                    WHERE version_id = ?
                      AND parse_status = 'parsed'
                      AND is_active = 1
                    """,
                    (now, max(0, int(chunk_count)), version_id),
                )

    def mark_index_failed(
        self,
        version_ids: Iterable[str],
        error: Exception,
    ) -> None:
        error_summary = _safe_error_summary(error)
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            for version_id in version_ids:
                connection.execute(
                    """
                    UPDATE rag_document_versions
                    SET index_status = 'index_failed', index_error = ?
                    WHERE version_id = ? AND is_active = 1
                    """,
                    (error_summary, str(version_id)),
                )

    def mark_inactive_removed(self) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                UPDATE rag_document_versions
                SET index_status = 'inactive', index_error = '',
                    indexed_at = ?, chunk_count = 0
                WHERE is_active = 0 AND index_status = 'pending_delete'
                """,
                (_utc_now(),),
            )
